import openpyxl
from selenium import webdriver
from bs4 import BeautifulSoup
import time

# Taking inputs
date = input('Enter date in the format(mm-dd-yyyy):')
inv = input('Enter invoice no(without #): ')
old = input('Enter the name of the excel file to be modified: ')

# Loading the file
wb = openpyxl.load_workbook(old)
ws = wb.worksheets[0]

# Deleting first 3 rows
ws.unmerge_cells('A1:C1')
ws.unmerge_cells('A3:F3')
ws.delete_rows(1, 3)

# Inserting 23 columns at the beginning
ws.insert_cols(1, 23)

# Calculating no of rows
rows = ws.max_row

# Shifting the columns
ws.move_range(cell_range='AR1:AR' + str(rows), cols=-43)
ws.cell(1, 1).value = 'Store'                             # Column A
ws.move_range(cell_range='AC1:AC' + str(rows), cols=-27)  # Column B
ws.move_range(cell_range='AK1:AK' + str(rows), cols=-34)  # Column C
ws.move_range(cell_range='AS1:AS' + str(rows), cols=-41)
ws.cell(1, 4).value = 'Customer Paid'                     # Column D
ws.move_range(cell_range='AJ1:AJ' + str(rows), cols=-31)
ws.cell(1, 5).value = 'Postage Paid'                      # Column E
ws.move_range(cell_range='AT1:AT' + str(rows), cols=-40)
ws.cell(1, 6).value = 'Profit/Unit'                       # Column F
ws.move_range(cell_range='AQ1:AQ' + str(rows), cols=-36)  # Column G
ws.move_range(cell_range='AB1:AB' + str(rows), cols=-20)  # Column H
ws.move_range(cell_range='AA1:AA' + str(rows), cols=-18)  # Column I
ws.move_range(cell_range='AG1:AG' + str(rows), cols=-23)  # Column J
ws.move_range(cell_range='AD1:AD' + str(rows), cols=-19)  # Column K
ws.move_range(cell_range='Z1:Z' + str(rows), cols=-14)    # Column L
ws.move_range(cell_range='AE1:AE' + str(rows), cols=-18)  # Column M
ws.move_range(cell_range='AF1:AF' + str(rows), cols=-18)  # Column N
ws.move_range(cell_range='X1:X' + str(rows), cols=-9)     # Column O
ws.move_range(cell_range='AH1:AH' + str(rows), cols=-18)  # Column P
ws.move_range(cell_range='Y1:Y' + str(rows), cols=-8)     # Column Q
ws.move_range(cell_range='AI1:AI' + str(rows), cols=-17)  # Column R
ws.move_range(cell_range='AL1:AL' + str(rows), cols=-19)  # Column S
ws.move_range(cell_range='AM1:AM' + str(rows), cols=-19)  # Column T
ws.move_range(cell_range='AN1:AN' + str(rows), cols=-19)  # Column U
ws.move_range(cell_range='AO1:AO' + str(rows), cols=-19)  # Column V
ws.move_range(cell_range='AP1:AP' + str(rows), cols=-19)  # Column W
ws.cell(1, 24).value = 'Already AUDITED or not'

# Deleting remaining columns
ws.delete_cols(47, 2)

# Loading the Webdriver
driver = webdriver.Chrome('chromedriver.exe')

# Log in
driver.get('https://ship6.shipstation.com/')
time.sleep(1)
input('Hit "ENTER" after the page is completely loaded: ')
driver.find_elements_by_class_name('advanced-search-text-6ODW4Fd')[0].click()

# Filling the columns
i = 2
while i <= rows:
    driver.find_elements_by_class_name('flex-input-3IbU2o7')[0].clear()
    driver.find_elements_by_class_name('flex-input-3IbU2o7')[0].send_keys(str(ws.cell(i, 2).value))
    driver.find_elements_by_class_name('flex-input-3IbU2o7')[1].clear()
    driver.find_elements_by_class_name('flex-input-3IbU2o7')[1].send_keys(str(ws.cell(i, 3).value))
    driver.find_elements_by_class_name('full-width-input-3F2-Knx')[0].clear()
    driver.find_elements_by_id('advanced-search-submit-button')[0].click()
    time.sleep(1)
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')
    try:
        # Column A
        store = soup.select('.message-Wl5p7I-')[0].getText()
        ws.cell(i, 1).value = str(store)
        # Column D
        total = soup.select('.currency-column-value-1VfNyxR')[0].getText()
        ws.cell(i, 4).value = float(str(total)[1:])
        # Column F
        if ws.cell(i, 13).value == 'COVID-19 Surcharge':
            ws.cell(i, 4).value = 0
        ws.cell(i, 6).value = '=IF(M' + str(i) + '="COVID-19 Surcharge","covid surcharge",D' + str(i) + '-E' + str(
            i) + ')'
        driver.find_elements_by_class_name('grid-rows-1E9Z-Ar')[1].click()
        driver.find_element_by_xpath("//div[contains(text(), 'Edit Tags')]").click()
        time.sleep(0.5)
        if driver.find_element_by_xpath('//button[@class="button-unstyled dropdown-menu-item-2NKgzjU dropdown-menu-item-1S4n1yj"][2]/div[1]/input[1]').is_selected():
            ws.cell(i, 24).value = 'Already AUDITED'
            ws.cell(i, 4).value = 0
            driver.find_element_by_xpath("//div[contains(text(), 'Edit Tags')]").click()
        else:
            driver.find_element_by_xpath("//button[@class='button-unstyled dropdown-menu-item-2NKgzjU dropdown-menu-item-1S4n1yj'][2]").click()
        time.sleep(2)
    except:
        i += 1
        continue
    i += 1

# Sum
ws.cell(rows + 1, 3).value = 'Totals'
ws.cell(rows + 1, 4).value = '=SUM(D2:D' + str(rows) + ')'
ws.cell(rows + 1, 5).value = '=SUM(E2:E' + str(rows) + ')'
ws.cell(rows + 1, 6).value = '=AVERAGE(F2:F' + str(rows) + ')'
ws.cell(rows + 2, 4).value = '=-E' + str(rows + 1)
ws.cell(rows + 4, 3).value = 'Gross profit'
ws.cell(rows + 4, 4).value = '=SUM(D' + str(rows + 1) + ':D' + str(rows + 2) + ')'
ws.cell(rows + 6, 3).value = 'Gross profit per parcel'
ws.cell(rows + 6, 4).value = '=D' + str(rows + 4) + '/77'
ws.cell(rows + 8, 3).value = 'MARGIN'
ws.cell(rows + 8, 4).value = '=D' + str(rows + 4) + '/D' + str(rows + 1)

# Setting each column width to 30
i = 'A'
while i != 'X':
    ws.column_dimensions[i].width = 30
    i = chr(ord(i) + 1)

# Save the file
wb.save('GLOBEGISTICS ' + date + ' - INVOICE ' + inv + '.xlsx')
